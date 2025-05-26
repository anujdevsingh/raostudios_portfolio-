from flask import render_template, jsonify, request, redirect, url_for, flash, session, send_file
from extensions import app, db
from models import Contact, Booking
from functools import wraps
import os
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Simple Admin Authentication
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Check if user is authenticated in session
        if not session.get('admin_authenticated'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/admin/login')
def admin_login():
    """Admin login page"""
    # If already authenticated, redirect to dashboard
    if session.get('admin_authenticated'):
        return redirect(url_for('admin_dashboard'))
        
    return render_template('admin_login.html')

@app.route('/admin/verify', methods=['POST'])
def admin_verify():
    """Verify admin password"""
    # Get admin password from environment or use default
    admin_password = os.environ.get('ADMIN_PASSWORD')
    
    # Get password from form
    password = request.form.get('password')
    
    if password == admin_password:
        # Set admin as authenticated in session
        session['admin_authenticated'] = True
        return redirect(url_for('admin_dashboard'))
    else:
        flash('Invalid password. Please try again.', 'danger')
        return redirect(url_for('admin_login'))

@app.route('/admin/logout')
def admin_logout():
    """Admin logout"""
    # Remove admin authentication from session
    session.pop('admin_authenticated', None)
    flash('You have been logged out successfully.', 'success')
    return redirect(url_for('admin_login'))

@app.route('/admin')
@admin_required
def admin_dashboard():
    """Admin dashboard showing database entries"""
    # Get latest contacts
    contacts = Contact.query.order_by(Contact.created_at.desc()).all()
    
    # Get latest bookings
    bookings = Booking.query.order_by(Booking.created_at.desc()).all()
    
    return render_template('admin_dashboard.html', 
                          contacts=contacts, 
                          bookings=bookings)

@app.route('/admin/export', methods=['GET'])
@admin_required
def admin_export():
    """Admin export data as JSON"""
    data_type = request.args.get('type', 'all')
    
    if data_type == 'contacts' or data_type == 'all':
        contacts = Contact.query.all()
        contacts_data = [{
            'id': c.id,
            'name': c.name,
            'email': c.email,
            'message': c.message,
            'created_at': c.created_at.isoformat() if c.created_at else None
        } for c in contacts]
    else:
        contacts_data = []
        
    if data_type == 'bookings' or data_type == 'all':
        bookings = Booking.query.all()
        bookings_data = [{
            'id': b.id,
            'name': b.name,
            'phone': b.phone,
            'event_type': b.event_type,
            'event_start_date': b.event_start_date.isoformat() if b.event_start_date else None,
            'event_end_date': b.event_end_date.isoformat() if b.event_end_date else None,
            'address': b.address,
            'notes': b.notes,
            'payment_id': b.payment_id,
            'payment_status': b.payment_status,
            'created_at': b.created_at.isoformat() if b.created_at else None
        } for b in bookings]
    else:
        bookings_data = []
    
    # Return data as JSON
    return jsonify({
        'contacts': contacts_data,
        'bookings': bookings_data
    })

@app.route('/admin/export/excel', methods=['GET'])
@admin_required
def admin_export_excel():
    """Admin export data as Excel file"""
    data_type = request.args.get('type', 'all')
    
    # Create a new workbook
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    if data_type == 'bookings' or data_type == 'all':
        # Create bookings sheet
        if data_type == 'all':
            ws_bookings = wb.active
            ws_bookings.title = "Bookings"
        else:
            ws_bookings = wb.active
            ws_bookings.title = "Bookings"
        
        # Bookings headers
        bookings_headers = [
            'ID', 'Name', 'Phone', 'Event Type', 'Start Date', 'End Date', 
            'Address', 'Notes', 'Payment ID', 'Payment Status', 'Created At'
        ]
        
        # Add headers
        for col, header in enumerate(bookings_headers, 1):
            cell = ws_bookings.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Get bookings data
        bookings = Booking.query.order_by(Booking.created_at.desc()).all()
        
        # Add data
        for row, booking in enumerate(bookings, 2):
            ws_bookings.cell(row=row, column=1, value=booking.id)
            ws_bookings.cell(row=row, column=2, value=booking.name)
            ws_bookings.cell(row=row, column=3, value=booking.phone)
            ws_bookings.cell(row=row, column=4, value=booking.event_type)
            ws_bookings.cell(row=row, column=5, value=booking.event_start_date.strftime('%Y-%m-%d') if booking.event_start_date else '')
            ws_bookings.cell(row=row, column=6, value=booking.event_end_date.strftime('%Y-%m-%d') if booking.event_end_date else '')
            ws_bookings.cell(row=row, column=7, value=booking.address)
            ws_bookings.cell(row=row, column=8, value=booking.notes)
            ws_bookings.cell(row=row, column=9, value=booking.payment_id)
            ws_bookings.cell(row=row, column=10, value=booking.payment_status)
            ws_bookings.cell(row=row, column=11, value=booking.created_at.strftime('%Y-%m-%d %H:%M:%S') if booking.created_at else '')
        
        # Auto-adjust column widths
        for col in range(1, len(bookings_headers) + 1):
            column_letter = get_column_letter(col)
            max_length = len(bookings_headers[col-1])
            for row in range(2, len(bookings) + 2):
                cell_value = str(ws_bookings.cell(row=row, column=col).value or '')
                max_length = max(max_length, len(cell_value))
            ws_bookings.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    if data_type == 'contacts' or data_type == 'all':
        # Create contacts sheet
        if data_type == 'all':
            ws_contacts = wb.create_sheet(title="Contacts")
        else:
            if data_type == 'contacts':
                ws_contacts = wb.active
                ws_contacts.title = "Contacts"
            else:
                ws_contacts = wb.create_sheet(title="Contacts")
        
        # Contacts headers
        contacts_headers = ['ID', 'Name', 'Email', 'Message', 'Created At']
        
        # Add headers
        for col, header in enumerate(contacts_headers, 1):
            cell = ws_contacts.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Get contacts data
        contacts = Contact.query.order_by(Contact.created_at.desc()).all()
        
        # Add data
        for row, contact in enumerate(contacts, 2):
            ws_contacts.cell(row=row, column=1, value=contact.id)
            ws_contacts.cell(row=row, column=2, value=contact.name)
            ws_contacts.cell(row=row, column=3, value=contact.email)
            ws_contacts.cell(row=row, column=4, value=contact.message)
            ws_contacts.cell(row=row, column=5, value=contact.created_at.strftime('%Y-%m-%d %H:%M:%S') if contact.created_at else '')
        
        # Auto-adjust column widths
        for col in range(1, len(contacts_headers) + 1):
            column_letter = get_column_letter(col)
            max_length = len(contacts_headers[col-1])
            for row in range(2, len(contacts) + 2):
                cell_value = str(ws_contacts.cell(row=row, column=col).value or '')
                max_length = max(max_length, len(cell_value))
            ws_contacts.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    if data_type == 'bookings':
        filename = f'rao_studios_bookings_{timestamp}.xlsx'
    elif data_type == 'contacts':
        filename = f'rao_studios_contacts_{timestamp}.xlsx'
    else:
        filename = f'rao_studios_all_data_{timestamp}.xlsx'
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    ) 